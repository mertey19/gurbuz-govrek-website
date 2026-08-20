"""
Tercih robotu veri çıkarımı: xlsx -> JSONL.

Çıktı `tmp/` altına yazılır ve .gitignore kapsamındadır. Veritabanına yükleme
`scripts/tercih/import.mjs`, veritabanından bağımsız salt okunur site verisi
üretme adımı ise `scripts/tercih/build-local-data.mjs` ile yapılır.

Kullanım:
    python scripts/tercih/extract.py "C:/yol/2026 TERCIH ROBOTU GURBUZ.xlsx"
"""

import json
import os
import re
import sys

import openpyxl

OUTPUT = os.path.join("tmp", "tercih-programs.jsonl")

# (sayfa adı, seviye, sütun indeksleri) — iki sayfanın sütun düzeni farklı.
SHEETS = {
    # 2026 "YENİ SON" düzeni. İki sayfa da aynı 16 sütunu taşır:
    # KOD, TÜR, İL, ÜNİVERSİTE, BÖLÜM, TÜR(puan), 2026 PUAN, SIRA DEĞİŞİMİ,
    # 2026/2025/2024/2023 SIRA, 2026/2025/2024/2023 KONT.
    # Fakülte, süre, koşullar, akademik kadro, akreditasyon ve TUS/DUS
    # sütunları bu sürümde yok; robot da bu alanları göstermiyor.
    "LİSANS": {
        "level": "lisans",
        "cols": {
            "program_code": 0, "kind": 1, "city": 2, "university": 3,
            "department": 4, "score_type": 5, "score": 6,
            "rank": 8, "rank_2025": 9, "rank_2024": 10, "rank_2023": 11,
            "quota": 12, "quota_2025": 13, "quota_2024": 14, "quota_2023": 15,
        },
    },
    # Önlisans programlarının puan türü her zaman TYT olduğu için sayfada
    # ayrı bir sütun yok; sabit veriliyor.
    "ÖNLİSANS": {
        "level": "onlisans",
        "default_score_type": "TYT",
        "cols": {
            "program_code": 0, "kind": 1, "city": 2, "university": 3,
            "department": 4, "score": 6,
            "rank": 8, "rank_2025": 9, "rank_2024": 10, "rank_2023": 11,
            "quota": 12, "quota_2025": 13, "quota_2024": 14, "quota_2023": 15,
        },
    },
}

# Sütun düzeni yıldan yıla değişiyor. Yanlış sütundan okumak sessizce bozuk veri
# üretir; bu yüzden başlık satırı beklenenle karşılaştırılır.
EXPECTED_HEADERS = [
    "KOD", "TÜR", "İL", "ÜNİVERSİTE", "BÖLÜM", "TÜR", "2026 PUAN",
    "SIRA DEĞİŞİMİ", "2026 SIRA", "2025 SIRA", "2024 SIRA", "2023 SIRA",
    "2026 KONT", "2025 KONT", "2024 KONT", "2023 KONT",
]


def check_headers(sheet, sheet_name):
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    actual = [clean_text(cell) or "" for cell in header][: len(EXPECTED_HEADERS)]
    if actual != EXPECTED_HEADERS:
        print(f"HATA: {sheet_name} sayfasının sütun düzeni beklenenden farklı.")
        for i, (bekl, olan) in enumerate(zip(EXPECTED_HEADERS, actual)):
            if bekl != olan:
                print(f"  sütun {i}: beklenen {bekl!r}, gelen {olan!r}")
        return False
    return True


def clean_text(value):
    if value is None:
        return None
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text or None


def to_int(value):
    """'12.500', '12500', 12500.0 -> 12500. Sayı değilse None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    digits = re.sub(r"[^\d]", "", str(value))
    return int(digits) if digits else None


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def main():
    if len(sys.argv) < 2:
        print("Kullanım: python scripts/tercih/extract.py <xlsx yolu>")
        return 1

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"Dosya bulunamadı: {path}")
        return 1

    os.makedirs("tmp", exist_ok=True)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    written = 0
    skipped = 0

    with open(OUTPUT, "w", encoding="utf-8") as out:
        for sheet_name, config in SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                print(f"UYARI: {sheet_name} sayfası bulunamadı, atlandı.")
                continue

            sheet = workbook[sheet_name]
            if not check_headers(sheet, sheet_name):
                return 1
            cols = config["cols"]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue

                def cell(key):
                    index = cols.get(key)
                    if index is None or index >= len(row):
                        return None
                    return row[index]

                # Sütun yoksa sayfa için tanımlı sabit tür kullanılır.
                score_type = clean_text(cell("score_type")) or config.get(
                    "default_score_type", ""
                )
                rank = to_int(cell("rank"))

                # Sıralaması veya puan türü olmayan satır robotta kullanılamaz.
                if not score_type or rank is None:
                    skipped += 1
                    continue

                record = {
                    "level": config["level"],
                    "program_code": clean_text(cell("program_code")),
                    "kind": clean_text(cell("kind")),
                    "city": clean_text(cell("city")),
                    "university": clean_text(cell("university")),
                    "department": clean_text(cell("department")),
                    "score_type": score_type.upper(),
                    # rank ve quota daima 2026 verisidir; tercih bu ikisiyle yapılır.
                    "rank": rank,
                    "score": to_float(cell("score")),
                    "quota": to_int(cell("quota")),
                    # Geçmiş yıllar trend gösterimi için tutulur; veri yoksa None.
                    "rank_2025": to_int(cell("rank_2025")),
                    "rank_2024": to_int(cell("rank_2024")),
                    "rank_2023": to_int(cell("rank_2023")),
                    "quota_2025": to_int(cell("quota_2025")),
                    "quota_2024": to_int(cell("quota_2024")),
                    "quota_2023": to_int(cell("quota_2023")),
                }

                out.write(json.dumps(record, ensure_ascii=False) + "\n")
                written += 1

    workbook.close()
    print(f"{written} program yazıldı -> {OUTPUT}")
    print(f"{skipped} satır atlandı (puan türü veya sıralama yok).")
    return 0


if __name__ == "__main__":
    sys.exit(main())
